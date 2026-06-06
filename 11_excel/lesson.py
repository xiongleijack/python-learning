"""
11 - Excel 办公自动化：按行数拆分大表

前置：07 异常与文件、04 函数。

依赖：
  pip install openpyxl

运行：
  python 11_excel/lesson.py

Java 对照：
  Apache POI / EasyExcel     openpyxl
  Files.walk                Path.glob
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook, load_workbook


def create_sample_excel(path: Path, *, data_rows: int = 25) -> None:
    """生成示例大表：1 行表头 + data_rows 行数据。"""
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "bonds"
    ws.append(["id", "bond_code", "yield"])
    for i in range(1, data_rows + 1):
        ws.append([i, f"24021{i % 10}.IB", round(2.0 + i * 0.01, 2)])
    wb.save(path)


def split_excel(source: Path, output_dir: Path, rows_per_file: int) -> list[Path]:
    """
    将 source 的第一个工作表按 rows_per_file 行数据拆成多个 xlsx。
    - 每个小文件都保留表头
    - 只计数据行，不含表头
    - 输出命名：{源文件名}_part001.xlsx, _part002.xlsx, ...
    """
    if rows_per_file <= 0:
        raise ValueError("rows_per_file 必须大于 0")

    output_dir.mkdir(parents=True, exist_ok=True)
    wb = load_workbook(source, read_only=True)
    ws = wb.active

    rows_iter = ws.iter_rows(values_only=True)
    header = next(rows_iter)
    if header is None:
        wb.close()
        raise ValueError("Excel 为空")

    written: list[Path] = []
    part_no = 1
    buffer: list[tuple] = []

    def flush(part: int, chunk: list[tuple]) -> Path | None:
        if not chunk:
            return None
        out = output_dir / f"{source.stem}_part{part:03d}.xlsx"
        out_wb = Workbook()
        out_ws = out_wb.active
        out_ws.append(list(header))
        for row in chunk:
            out_ws.append(list(row))
        out_wb.save(out)
        return out

    for row in rows_iter:
        buffer.append(row)
        if len(buffer) >= rows_per_file:
            path = flush(part_no, buffer)
            if path:
                written.append(path)
            buffer = []
            part_no += 1

    path = flush(part_no, buffer)
    if path:
        written.append(path)

    wb.close()
    return written


def main() -> None:
    base = Path(__file__).parent
    sample = base / "samples" / "bonds_large.xlsx"
    out_dir = base / "output" / "demo"

    create_sample_excel(sample, data_rows=25)
    files = split_excel(sample, out_dir, rows_per_file=10)

    print(f"源文件: {sample}（1 表头 + 25 行数据）")
    print(f"每文件最多 10 行数据 → 共 {len(files)} 个小文件:")
    for f in files:
        print(f"  {f.name}")


if __name__ == "__main__":
    main()
