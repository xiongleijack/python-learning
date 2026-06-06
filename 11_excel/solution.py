"""practice.py 参考答案"""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook, load_workbook


def split_excel(source: Path, output_dir: Path, rows_per_file: int) -> list[Path]:
    if rows_per_file <= 0:
        raise ValueError("rows_per_file 必须大于 0")

    output_dir.mkdir(parents=True, exist_ok=True)
    wb = load_workbook(source, read_only=True)
    ws = wb.active

    rows_iter = ws.iter_rows(values_only=True)
    header = next(rows_iter)
    if header is None:
        wb.close()
        raise ValueError("Excel 为空")

    written: list[Path] = []
    part_no = 1
    buffer: list[tuple] = []

    def flush(part: int, chunk: list[tuple]) -> Path | None:
        if not chunk:
            return None
        out = output_dir / f"{source.stem}_part{part:03d}.xlsx"
        out_wb = Workbook()
        out_ws = out_wb.active
        out_ws.append(list(header))
        for row in chunk:
            out_ws.append(list(row))
        out_wb.save(out)
        return out

    for row in rows_iter:
        buffer.append(row)
        if len(buffer) >= rows_per_file:
            path = flush(part_no, buffer)
            if path:
                written.append(path)
            buffer = []
            part_no += 1

    path = flush(part_no, buffer)
    if path:
        written.append(path)

    wb.close()
    return written
